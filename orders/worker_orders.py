import openpyxl
import datetime


DATA = openpyxl.load_workbook('data/data.xlsx')
WORKERS_SHEET = DATA['workers']
ORDERS_SHEET = DATA['orders']


def worker_get_product(product):
    product_name = str(product).split()
    for i in product_name:
        if i.find('_') == 0:
            return i[1:]


def worker_get_all_products(num_order):
    products = ''
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order):
            for i in str(ORDERS_SHEET[f'G{i+1}'].value).split():
                if i.find('-') == 0:
                    if i == '-мороженое':
                        products += i[1:] + ' '
                    else:
                        products += i[1:] + ', '
                elif i.find(':') == 0:
                    products += 'вкус: ' + i[1:] + ' '
                elif i.find('+') == 0:
                    products += i + ', '
    return str(products)


def worker_create_new_order(worker_id):
    max_rw_workers = WORKERS_SHEET.max_row
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_workers):
        if str(WORKERS_SHEET[f'C{i+1}'].value) == str(worker_id):
            # +1 чтобы захватывалась пустая ячейка
            for i in range(max_rw_orders + 1):
                if ORDERS_SHEET[f'A{i+1}'].value == None:
                    ORDERS_SHEET[f'A{i+1}'] = str(worker_id)
                    ORDERS_SHEET[f'B{i+1}'] = str(int(
                        ORDERS_SHEET[f'A{i+1}'].row - 1))
                    ORDERS_SHEET[f'C{i+1}'] = str('Создается')
                    ORDERS_SHEET[f'F{i+1}'] = str(datetime.datetime.now())
                    DATA.save('data/data.xlsx')
                    DATA.close()
                    return str(int(ORDERS_SHEET[f'A{i+1}'].row - 1))


def worker_get_count_created_order(worker_id):
    num_order = []
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'A{i+1}'].value) == str(worker_id) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            num_order.append(str(int(ORDERS_SHEET[f'B{i+1}'].value)))
    return str(max(num_order))


def worker_get_count_all_orders():
    # Минусуем чтобы последняя строка не попадала
    # Иначе будет на выходе None в функции check_actua_orders
    return int(ORDERS_SHEET.max_row) - 1


def worker_append_product_to_order(product, num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            if ORDERS_SHEET[f'G{i+1}'].value != None:
                ORDERS_SHEET[f'G{i+1}'] = f'{ORDERS_SHEET[f"G{i+1}"].value} -{product}'
            elif ORDERS_SHEET[f'G{i+1}'].value == None:
                ORDERS_SHEET[f'G{i+1}'] = f'-{product}'
    DATA.save('data/data.xlsx')
    DATA.close()


def worker_append_additional_product_to_order(product, num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            if ORDERS_SHEET[f'G{i+1}'].value != None:
                ORDERS_SHEET[f'G{i+1}'] = f'{ORDERS_SHEET[f"G{i+1}"].value} +{product}'
    DATA.save('data/data.xlsx')
    DATA.close()


def worker_complete_create_order(num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            ORDERS_SHEET[f'C{i+1}'] = str('Выполняется')
            ORDERS_SHEET[f'F{i+1}'] = str(datetime.datetime.now())
    DATA.save('data/data.xlsx')
    DATA.close()


def worker_complete_order():
    pass


def worker_remove_order():
    pass


def worker_check_completed_orders():
    pass


def worker_check_actual_orders(num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Выполняется'):
            order_info = f'Заказ №{num_order}: {worker_get_all_products(num_order)} ЦЕНА: 0'
            return order_info


def worker_append_cash_order():
    pass
