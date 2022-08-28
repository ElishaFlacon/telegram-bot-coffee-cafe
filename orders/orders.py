import openpyxl
import datetime


DATA = openpyxl.load_workbook('data/data.xlsx')
WORKERS_SHEET = DATA['workers']
ORDERS_SHEET = DATA['orders']


# Создаем новый заказ
def create_new_order(worker_id):
    max_rw_workers = WORKERS_SHEET.max_row
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_workers):
        if str(WORKERS_SHEET[f'C{i+1}'].value) == str(worker_id):
            # +1 чтобы захватывалась пустая ячейка
            for i in range(max_rw_orders + 1):
                if ORDERS_SHEET[f'A{i+1}'].value == None:
                    ORDERS_SHEET[f'A{i+1}'] = str(worker_id)
                    ORDERS_SHEET[f'B{i+1}'] = str(int(
                        ORDERS_SHEET[f'A{i+1}'].row - 1))
                    ORDERS_SHEET[f'C{i+1}'] = str('Создается')
                    ORDERS_SHEET[f'F{i+1}'] = str(datetime.datetime.now())
                    DATA.save('data/data.xlsx')
                    DATA.close()
                    return str(int(ORDERS_SHEET[f'A{i+1}'].row - 1))


# Завершаем создание заказа
def complete_create_order(num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            ORDERS_SHEET[f'C{i+1}'] = str('Выполняется')
            ORDERS_SHEET[f'F{i+1}'] = str(datetime.datetime.now())
    DATA.save('data/data.xlsx')
    DATA.close()


# Получаем количество всех заказов
def get_count_all_orders():
    # Минусуем чтобы последняя строка не попадала
    # Иначе будет на выходе None в функции check_actua_orders
    return int(ORDERS_SHEET.max_row) - 1


# Получаем количество создающихся заказов
def get_count_being_created_order(worker_id):
    num_order = []
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'A{i+1}'].value) == str(worker_id) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            num_order.append(str(int(ORDERS_SHEET[f'B{i+1}'].value)))
    return str(max(num_order))


# Выполняем заказ
def complete_order():
    pass


# Отменяем заказ
def remove_order():
    pass


# Смотрим выполняющиеся заказы
def check_running_orders(num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Выполняется'):
            order_info = f'Заказ №{num_order}: {get_all_products(num_order)} ЦЕНА: 0'
            return order_info


# Смотрим выполненные заказы
def check_completed_orders():
    pass


# Получаем продукт
def get_product(product):
    product_name = str(product).split()
    for i in product_name:
        if i.find('_') == 0:
            return i[1:]


# Получаем все продукты из заказа
def get_all_products(num_order):
    products = ''
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order):
            for i in str(ORDERS_SHEET[f'G{i+1}'].value).split():
                if i.find('-') == 0:
                    if i == '-мороженое':
                        products += i[1:] + ' '
                    else:
                        products += i[1:] + ', '
                elif i.find(':') == 0:
                    products += 'вкус: ' + i[1:] + ' '
                elif i.find('+') == 0:
                    products += i + ', '
    return str(products)


# Добовляем продукты в заказ
def append_product_to_order(product, num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        print(
            f"{str(ORDERS_SHEET[f'B{i+1}'].value)} == {str(num_order)} ----- {str(ORDERS_SHEET[f'C{i+1}'].value)} == {str('Создается')}")
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            if ORDERS_SHEET[f'G{i+1}'].value != None:
                ORDERS_SHEET[f'G{i+1}'] = f'{ORDERS_SHEET[f"G{i+1}"].value} -{product}'
                print(f'{ORDERS_SHEET[f"G{i+1}"].value} -{product}')
            elif ORDERS_SHEET[f'G{i+1}'].value == None:
                ORDERS_SHEET[f'G{i+1}'] = f'-{product}'
                print(f'-{product}')
    DATA.save('data/data.xlsx')
    DATA.close()


# Добовляем дополнительные продукты в заказ, типо вкус мороженого
def append_additional_product_to_order(product, num_order):
    max_rw_orders = ORDERS_SHEET.max_row
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order) and str(ORDERS_SHEET[f'C{i+1}'].value) == str('Создается'):
            if ORDERS_SHEET[f'G{i+1}'].value != None:
                ORDERS_SHEET[f'G{i+1}'] = f'{ORDERS_SHEET[f"G{i+1}"].value} +{product}'
    DATA.save('data/data.xlsx')
    DATA.close()


# Добовляем цену продукта в заказ
def append_product_price_to_order():
    pass
