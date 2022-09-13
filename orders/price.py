from re import A
import openpyxl
import time

DATA = openpyxl.load_workbook('data/data.xlsx')
WORKERS_SHEET = DATA['workers']
ADMINS_SHEET = DATA['admins']
SESSIONS_SHEET = DATA['sessions']
ORDERS_SHEET = DATA['orders']
PRODUCTS_SHEET = DATA['products']
CASH_SHEET = DATA['cash']


def save_data():
    DATA.save('data/data.xlsx')
    DATA.close()


# !функ ставим в функцию выполненного заказа перед тем как функция ваполнен произойдет
# Получаем цену заказа, который Создается
def get_process_order_price(num_order):
    max_rw_products = PRODUCTS_SHEET.max_row
    max_rw_orders = ORDERS_SHEET.max_row
    check_additions = 0
    order_price = 0
    for i in range(max_rw_orders):
        if str(ORDERS_SHEET[f'B{i+1}'].value) == str(num_order):
            # Пробел в сплите, после ; нужен, просто необходим, наверно))
            # [:-1] нужен, зачем? а потому что в конце массива появлялась пустая строка
            # через иф не убиралась, так что вырежем так))
            for j in ORDERS_SHEET[f'G{i+1}'].value.split('; ')[:-1]:
                for l in range(max_rw_products):
                    # Тут проверка на продукт и вкус
                    if j.lower().find(PRODUCTS_SHEET[f'A{l+1}'].value) == 0:
                        a = PRODUCTS_SHEET[f'B{l+1}'].value
                        order_price += int(a)
                    # А тут проверка на посыпку и топинг
                    # >=5 потому что да
                    # А вообще функция финд находит кол-во символов до указанного символа
                    # И соответственно возвращает их
                    # А до + который является началом для посыпок достаточно далеко, а топинг еще дальше
                    # Но минимально лучше оставить 5
                    elif j.lower().find(PRODUCTS_SHEET[f'A{l+1}'].value) >= 5:
                        for i in j.lower().split():
                            if i.find('+') >= 0:
                                # Проверка нужна, потому что первая посыпка бесплатная
                                if check_additions == 1:
                                    # Тут мне было в падлу нормально делать поиск по +
                                    #  Плюс у нас это топинг, вот
                                    #  Так что столбец у нас, такой, константа
                                    # Ну и там ниже так же
                                    b = PRODUCTS_SHEET[f'B{2}'].value
                                    order_price += int(b)
                                check_additions = 1
                            # Топинг хоть и бесплатный, но а вдруг... все бывает))
                            elif i.find('=') >= 0:
                                b = PRODUCTS_SHEET[f'B{3}'].value
                                order_price += int(b)
    print(order_price, '---')


get_process_order_price(2)
