import openpyxl


DATA = openpyxl.load_workbook('data/data.xlsx')
WORKERS_SHEET = DATA['workers']


def worker_session_status(worker_id):
    max_rw = WORKERS_SHEET.max_row
    for i in range(max_rw):
        if str(WORKERS_SHEET[f'C{i+1}'].value) == str(worker_id) and str(WORKERS_SHEET[f'D{i+1}'].value) == str('Да'):
            return True
        elif str(WORKERS_SHEET[f'C{i+1}'].value) == str(worker_id) and str(WORKERS_SHEET[f'D{i+1}'].value) == str('Нет'):
            return False


def worker_start_session(worker_id):
    max_rw = WORKERS_SHEET.max_row
    for i in range(max_rw):
        if str(WORKERS_SHEET[f'C{i+1}'].value) == str(worker_id):
            WORKERS_SHEET[f'D{i+1}'] = 'Да'
    DATA.save('data/data.xlsx')
    DATA.close()


def worker_end_session(worker_id):
    max_rw = WORKERS_SHEET.max_row
    for i in range(max_rw):
        if str(WORKERS_SHEET[f'C{i+1}'].value) == str(worker_id):
            WORKERS_SHEET[f'D{i+1}'] = 'Нет'
    DATA.save('data/data.xlsx')
    DATA.close()
